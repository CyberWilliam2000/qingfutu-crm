from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font
import os
import json
import re
import os
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, static_folder=BASE_DIR, static_url_path='')
CORS(app)

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, '客户追踪看板.html')

@app.route('/客户追踪看板.html')
def dashboard():
    return send_from_directory(BASE_DIR, '客户追踪看板.html')

EXCEL_PATH = os.path.join(BASE_DIR, '客户追踪看板.xlsx')
HEADER_ROW = 4  # Row 4 (1-indexed) = Row index 3 in openpyxl

COL_MAP = {
    'id': 1,       # A: ID
    'name': 2,     # B: 客户名
    'biz': 3,      # C: 业务线
    'proj': 4,     # D: 项目
    'amount': 5,   # E: 预估金额(CNY)
    'currency': 6, # F: 币种
    'stage': 7,    # G: 当前阶段
    'owner': 8,    # H: 负责人
    'entry': 9,    # I: 进入日
    'next': 10,    # J: 下次跟进
    'note': 11,    # K: 状态/备注
    'referrer': 12,# L: 介绍人/来源
    'action': 13,  # M: 操作动作
}

def read_excel_customers():
    """Read all customers from Excel 客户列表 sheet"""
    if not os.path.exists(EXCEL_PATH):
        return []
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['客户列表']
    
    customers = []
    for row in ws.iter_rows(min_row=HEADER_ROW, max_col=13):
        # row is 0-indexed tuple
        cid = row[0].value
        if not cid or str(cid).strip() in ['', 'ID', '合计']:
            continue
        
        name = str(row[1].value or '')
        biz_raw = str(row[2].value or '')
        # biz can be comma-separated or single
        biz = [b.strip() for b in biz_raw.replace('，', ',').split(',') if b.strip()]
        if not biz:
            biz = ['调研']
        
        proj = str(row[3].value or '')
        amount = str(row[4].value or '')
        currency = str(row[5].value or 'CNY')
        stage = str(row[6].value or '')
        owner = str(row[7].value or '')
        entry = str(row[8].value or '')
        next_date = str(row[9].value or '')
        note = str(row[10].value or '')
        referrer = str(row[11].value or '')
        
        # Parse next_date: if it looks like a datetime, extract date part
        if next_date and next_date != 'None':
            try:
                dt = datetime.strptime(next_date[:10], '%Y-%m-%d')
                next_date = dt.strftime('%Y-%m-%d')
            except:
                pass
        
        customers.append({
            'id': str(cid),
            'name': name,
            'biz': biz,
            'proj': proj,
            'amount': amount,
            'currency': currency,
            'stage': stage,
            'owner': owner,
            'entry': entry,
            'next': next_date if next_date != 'None' else '',
            'note': note,
            'referrer': referrer,
            'isNew': False,
        })
    
    wb.close()
    return customers


def write_excel_customers(customers):
    """Write customers back to Excel 客户列表 sheet, preserving all formatting"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['客户列表']
    
    # Build a map of id -> customer for quick lookup
    cust_map = {}
    for c in customers:
        cust_map[c['id']] = c
    
    # Scan existing rows, update matching IDs
    existing_ids = set()
    data_start_row = HEADER_ROW  # Row 4
    last_data_row = data_start_row
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        cid = ws.cell(row=row_idx, column=1).value
        if not cid or str(cid).strip() in ['', 'ID', '合计']:
            continue
        
        cid_str = str(cid)
        existing_ids.add(cid_str)
        last_data_row = row_idx
        
        if cid_str in cust_map:
            c = cust_map[cid_str]
            ws.cell(row=row_idx, column=COL_MAP['name']).value = c['name']
            ws.cell(row=row_idx, column=COL_MAP['biz']).value = ', '.join(c.get('biz', ['调研']))
            ws.cell(row=row_idx, column=COL_MAP['proj']).value = c.get('proj', '')
            ws.cell(row=row_idx, column=COL_MAP['stage']).value = c.get('stage', '')
            ws.cell(row=row_idx, column=COL_MAP['owner']).value = c.get('owner', '')
            next_val = c.get('next', '')
            if next_val:
                ws.cell(row=row_idx, column=COL_MAP['next']).value = next_val
            else:
                ws.cell(row=row_idx, column=COL_MAP['next']).value = ''
            ws.cell(row=row_idx, column=COL_MAP['note']).value = c.get('note', '')
            ws.cell(row=row_idx, column=COL_MAP['referrer']).value = c.get('referrer', '')
    
    # Add new customers that don't exist in Excel
    new_customers = [c for c in customers if c['id'] not in existing_ids]
    if new_customers:
        insert_row = last_data_row + 1
        for c in new_customers:
            ws.insert_rows(insert_row)
            ws.cell(row=insert_row, column=COL_MAP['id']).value = c['id']
            ws.cell(row=insert_row, column=COL_MAP['name']).value = c['name']
            ws.cell(row=insert_row, column=COL_MAP['biz']).value = ', '.join(c.get('biz', ['调研']))
            ws.cell(row=insert_row, column=COL_MAP['proj']).value = c.get('proj', '')
            ws.cell(row=insert_row, column=COL_MAP['amount']).value = c.get('amount', '')
            ws.cell(row=insert_row, column=COL_MAP['currency']).value = c.get('currency', 'CNY')
            ws.cell(row=insert_row, column=COL_MAP['stage']).value = c.get('stage', '')
            ws.cell(row=insert_row, column=COL_MAP['owner']).value = c.get('owner', '')
            ws.cell(row=insert_row, column=COL_MAP['entry']).value = c.get('entry', '')
            ws.cell(row=insert_row, column=COL_MAP['next']).value = c.get('next', '')
            ws.cell(row=insert_row, column=COL_MAP['note']).value = c.get('note', '')
            ws.cell(row=insert_row, column=COL_MAP['referrer']).value = c.get('referrer', '')
            insert_row += 1
    
    wb.save(EXCEL_PATH)
    wb.close()
    return True


@app.route('/api/customers', methods=['GET'])
def get_customers():
    try:
        customers = read_excel_customers()
        return jsonify({'success': True, 'customers': customers})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/customers/sync', methods=['POST'])
def sync_customers():
    try:
        data = request.get_json()
        customers = data.get('customers', [])
        write_excel_customers(customers)
        return jsonify({'success': True, 'message': f'已同步 {len(customers)} 条客户数据到 Excel'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'excel_exists': os.path.exists(EXCEL_PATH)})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
